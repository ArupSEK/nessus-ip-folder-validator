from __future__ import annotations

import csv
import ipaddress
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.app.models.scan import ScanRecord
from backend.app.schemas.ip_search import IpSearchMatch, IpSearchResponse, IpSearchResultItem
from backend.app.services.auth import ensure_utc
from ip_utils import normalize_ip

CIDR_EXPANSION_LIMIT = 256


class IpSearchError(ValueError):
    pass


def _collect_text_cells(text: str) -> list[str]:
    rows: list[str] = []
    reader = csv.reader(StringIO(text))
    for row in reader:
        for cell in row:
            stripped = cell.strip()
            if stripped:
                rows.append(stripped)
    return rows


def parse_uploaded_entries(filename: str, payload: bytes) -> list[str]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".txt":
        return [line.strip() for line in payload.decode("utf-8", errors="ignore").splitlines() if line.strip()]
    if suffix == ".csv":
        return _collect_text_cells(payload.decode("utf-8", errors="ignore"))
    if suffix == ".xlsx":
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        values: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    if text:
                        values.append(text)
        return values
    raise IpSearchError("Unsupported file type. Use .txt, .csv or .xlsx.")


def _expand_if_cidr(entry: str, *, expand_cidr: bool) -> list[str]:
    if "/" not in entry:
        return [entry]
    if not expand_cidr:
        return [entry]
    try:
        network = ipaddress.ip_network(entry, strict=False)
    except ValueError:
        return [entry]
    hosts = list(network.hosts()) if network.num_addresses > 2 else list(network)
    if len(hosts) > CIDR_EXPANSION_LIMIT:
        raise IpSearchError("CIDR expansion exceeds the current safety limit.")
    return [str(host) for host in hosts]


def normalize_search_entries(entries: list[str], *, expand_cidr: bool) -> tuple[list[str], list[str]]:
    unique: list[str] = []
    invalid: list[str] = []
    seen: set[str] = set()
    for raw in entries:
        stripped = raw.strip()
        if not stripped:
            continue
        expanded_items = _expand_if_cidr(stripped, expand_cidr=expand_cidr)
        for item in expanded_items:
            normalized = normalize_ip(item)
            if normalized is None:
                invalid.append(raw)
                continue
            if normalized not in seen:
                unique.append(normalized)
                seen.add(normalized)
    return unique, invalid


def _scan_targets(scan: ScanRecord) -> list[str]:
    return [item.strip() for item in scan.targets_text.split(",") if item.strip()]


def _target_matches(query_ip: str, target: str) -> bool:
    normalized = normalize_ip(target)
    if normalized == query_ip:
        return True
    if "/" in target:
        try:
            return ipaddress.ip_address(query_ip) in ipaddress.ip_network(target, strict=False)
        except ValueError:
            return False
    return False


def run_ip_search(db: Session, entries: list[str], *, expand_cidr: bool = False) -> IpSearchResponse:
    normalized_entries, invalid = normalize_search_entries(entries, expand_cidr=expand_cidr)
    scans = db.scalars(select(ScanRecord).where(ScanRecord.deleted_at.is_(None)).order_by(ScanRecord.name)).all()
    results: list[IpSearchResultItem] = []
    for query_ip in normalized_entries:
        matches: list[IpSearchMatch] = []
        for scan in scans:
            for target in _scan_targets(scan):
                if _target_matches(query_ip, target):
                    last_scan = ensure_utc(scan.last_completion_at) or ensure_utc(scan.last_launch_at)
                    matches.append(
                        IpSearchMatch(
                            query=query_ip,
                            normalized_ip=query_ip,
                            folder_name=scan.folder_name,
                            scan_name=scan.name,
                            scan_status=scan.status,
                            reachability="unknown",
                            authentication_status="unknown",
                            credentialed_checks_status="unknown",
                            last_scan_date=last_scan.isoformat() if last_scan else None,
                        )
                    )
                    break
        results.append(IpSearchResultItem(query=query_ip, normalized_ip=query_ip, matches=matches))
    return IpSearchResponse(
        total_inputs=len(entries),
        unique_inputs=len(normalized_entries),
        invalid_inputs=invalid,
        results=results,
    )
