from __future__ import annotations

import json
from io import BytesIO
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from backend.app.models.asset import AssetRecord
from backend.app.models.auth import AuditEvent
from backend.app.models.finding import FindingRecord
from backend.app.models.import_job import ImportJob
from backend.app.models.scan import ScanRecord


def _seed_job(db_session, *, scan_name: str, template_uuid: str, status_meta: dict[str, str] | None = None):
    scan = ScanRecord(
        nessus_scan_id=str(uuid4()),
        nessus_uuid=str(uuid4()),
        name=scan_name,
        template_uuid=template_uuid,
        status="completed",
    )
    db_session.add(scan)
    db_session.flush()
    job = ImportJob(scan_record_id=scan.id, job_scope_key=str(uuid4()), status="completed", progress_percent=100)
    db_session.add(job)
    db_session.flush()
    asset = AssetRecord(
        stable_asset_key=f"{scan_name.lower()}-asset",
        source_import_job_id=job.id,
        source_scan_record_id=scan.id,
        hostname=scan_name.lower(),
        ipv4_address="10.0.0.1",
        raw_metadata=json.dumps(status_meta or {}, sort_keys=True),
    )
    db_session.add(asset)
    db_session.flush()
    return scan, job, asset


async def _login(client, username: str, password: str) -> str:
    response = await client.post("/api/v1/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200
    return response.json()["csrf_token"]


@pytest.mark.anyio
async def test_dashboard_summary_and_drilldown(client, admin_user, db_session) -> None:
    _, previous_job, previous_asset = _seed_job(db_session, scan_name="Dash", template_uuid="tmpl-a")
    _, latest_job, latest_asset = _seed_job(
        db_session,
        scan_name="Dash",
        template_uuid="tmpl-a",
        status_meta={"reachability_status": "unreachable", "authentication_status": "failed", "credentialed_checks_status": "failed"},
    )
    db_session.add(FindingRecord(finding_key="dash-asset:1001:443:tcp", source_import_job_id=previous_job.id, source_scan_record_id=previous_job.scan_record_id, asset_record_id=previous_asset.id, plugin_id=1001, plugin_name="Old", severity=4, port=443, protocol="tcp"))
    db_session.add(FindingRecord(finding_key="dash-asset:2002:80:tcp", source_import_job_id=latest_job.id, source_scan_record_id=latest_job.scan_record_id, asset_record_id=latest_asset.id, plugin_id=2002, plugin_name="New", severity=2, port=80, protocol="tcp"))
    db_session.commit()

    csrf = await _login(client, "admin", "StrongPass123!")
    comparison = await client.post("/api/v1/comparisons/run", headers={"X-CSRF-Token": csrf}, json={"previous_import_job_id": previous_job.id, "latest_import_job_id": latest_job.id})
    assert comparison.status_code == 200
    comparison_run_id = comparison.json()["id"]

    summary = await client.get("/api/v1/dashboard/summary", params={"comparison_run_id": comparison_run_id})
    assert summary.status_code == 200
    body = summary.json()
    assert body["comparison_run_id"] == comparison_run_id
    assert body["previous_total"] == 1
    assert body["latest_total"] == 1
    assert body["new"] == 1
    assert body["not_validated"] == 1
    assert body["asset_coverage"]["unreachable_assets"] == 1
    assert body["asset_coverage"]["authentication_failed"] == 1
    assert body["asset_coverage"]["credentialed_checks_failed"] == 1

    filtered = await client.get("/api/v1/dashboard/findings", params={"comparison_run_id": comparison_run_id, "lifecycle_status": "New", "search": "2002"})
    assert filtered.status_code == 200
    filtered_body = filtered.json()
    assert filtered_body["total"] == 1
    assert filtered_body["findings"][0]["plugin_id"] == 2002
    assert filtered_body["findings"][0]["lifecycle_status"] == "New"


@pytest.mark.anyio
async def test_report_export_permissions_and_formula_safety(client, admin_user, readonly_user, db_session) -> None:
    _, previous_job, previous_asset = _seed_job(db_session, scan_name="Risk", template_uuid="tmpl-a")
    _, latest_job, latest_asset = _seed_job(db_session, scan_name="Risk", template_uuid="tmpl-a")
    db_session.add(FindingRecord(finding_key="=cmd|' /C calc'!A0", source_import_job_id=previous_job.id, source_scan_record_id=previous_job.scan_record_id, asset_record_id=previous_asset.id, plugin_id=1001, plugin_name="@formula", severity=3, port=443, protocol="tcp"))
    db_session.add(FindingRecord(finding_key="=cmd|' /C calc'!A0", source_import_job_id=latest_job.id, source_scan_record_id=latest_job.scan_record_id, asset_record_id=latest_asset.id, plugin_id=1001, plugin_name="@formula", severity=3, port=443, protocol="tcp"))
    db_session.add(AuditEvent(action="scans.delete", object_type="scan", object_id="1", object_name="=DELETE-ME", result="success"))
    db_session.commit()

    admin_csrf = await _login(client, "admin", "StrongPass123!")
    comparison = await client.post("/api/v1/comparisons/run", headers={"X-CSRF-Token": admin_csrf}, json={"previous_import_job_id": previous_job.id, "latest_import_job_id": latest_job.id})
    assert comparison.status_code == 200
    comparison_run_id = comparison.json()["id"]

    readonly_csrf = await _login(client, "viewer", "StrongPass123!")
    denied = await client.get("/api/v1/reports/export", params={"report_type": "scan_comparison", "export_format": "csv", "comparison_run_id": comparison_run_id}, headers={"X-CSRF-Token": readonly_csrf})
    assert denied.status_code == 403

    admin_csrf = await _login(client, "admin", "StrongPass123!")
    csv_export = await client.get("/api/v1/reports/export", params={"report_type": "scan_comparison", "export_format": "csv", "comparison_run_id": comparison_run_id}, headers={"X-CSRF-Token": admin_csrf})
    assert csv_export.status_code == 200
    assert "text/csv" in csv_export.headers["content-type"]
    csv_text = csv_export.text
    assert "'=cmd|' /C calc'!A0" in csv_text

    xlsx_export = await client.get("/api/v1/reports/export", params={"report_type": "deleted_objects_audit", "export_format": "xlsx"}, headers={"X-CSRF-Token": admin_csrf})
    assert xlsx_export.status_code == 200
    workbook = load_workbook(filename=BytesIO(xlsx_export.content))
    sheet = workbook.active
    values = [cell for row in sheet.iter_rows(values_only=True) for cell in row if isinstance(cell, str)]
    assert "'=DELETE-ME" in values
